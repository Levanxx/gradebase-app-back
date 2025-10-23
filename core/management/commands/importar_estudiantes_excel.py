from django.core.management.base import BaseCommand, CommandParser
from django.db import transaction, IntegrityError
from django.contrib.auth.models import User
from openpyxl import load_workbook

from core.models import Curso, Seccion, Estudiante, Nota

CODE_MAP = {
    "full stack": "FS",
    "desarrollo de aplicaciones moviles": "DAM",
    "herramientas de desarrollo": "HD",
    "inteligencia artificial": "IA",
    "business intelligence": "BI",
    "base de datos": "BD",
    "data visualitation": "DV",
    "python para ciencia de datos": "PCD",
    "big data": "BIGD",
    "desarrollo de videojuegos": "DVJ",
}

def initials_code(name: str) -> str:
    parts = [p for p in name.strip().split() if p]
    code = ''.join(p[0] for p in parts).upper()
    return code[:10] or "CURSO"

class Command(BaseCommand):
    help = "Importa estudiantes/cursos/secciones desde un Excel. Columnas: curso, codigo, nombre, apellido, email"

    def add_arguments(self, parser: CommandParser) -> None:
        parser.add_argument("excel_path", help="Ruta al .xlsx")
        parser.add_argument("--seccion", default="A", help="Nombre de la sección (default: A)")
        parser.add_argument("--profesor", default=None, help="Username del profesor para asignar a las secciones")
        parser.add_argument("--con-nota", action="store_true", help="Crear nota semilla (0) por alumno")

    @transaction.atomic
    def handle(self, *args, **opts):
        path = opts["excel_path"]
        seccion_nombre = opts["seccion"]
        profesor_username = opts.get("profesor")
        crear_nota = bool(opts.get("con-nota"))

        profesor = None
        if profesor_username:
            profesor = User.objects.filter(username=profesor_username).first()
            if not profesor:
                self.stderr.write(self.style.WARNING(
                    f"⚠ Profesor '{profesor_username}' no existe. Secciones sin profesor."
                ))

        wb = load_workbook(filename=path, data_only=True)

        total_rows = created_cursos = created_secciones = created_estudiantes = created_notas = 0

        def header_map(sheet):
            hdr = {}
            first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            for i, cell in enumerate(first, start=1):
                if cell:
                    hdr[str(cell).strip().lower()] = i
            return hdr

        for sheet in wb.worksheets:
            headers = header_map(sheet)
            curso_from_sheet = sheet.title

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue

                def val(col, default=None):
                    idx = headers.get(col)
                    if not idx: return default
                    v = row[idx-1]
                    return v if v is not None else default

                curso_nombre = str(val("curso", curso_from_sheet)).strip()
                codigo = str(val("codigo", "")).strip()
                nombre = str(val("nombre", "")).strip()
                apellido = str(val("apellido", "")).strip()
                email = str(val("email", "")).strip().lower()

                if not (curso_nombre and codigo and nombre and apellido):
                    continue

                if not email:
                    email = f"{codigo.lower()}@utepe.edu.pe"

                # ----- Curso -----
                codigo_curso = CODE_MAP.get(curso_nombre.lower(), initials_code(curso_nombre))
                curso, c_created = Curso.objects.get_or_create(
                    codigo=codigo_curso,
                    defaults={"nombre": curso_nombre}
                )
                if c_created:
                    created_cursos += 1
                elif curso.nombre != curso_nombre:
                    curso.nombre = curso_nombre
                    curso.save(update_fields=["nombre"])

                # ----- Sección -----
                seccion_defaults = {"profesor": profesor} if profesor else {}
                seccion, s_created = Seccion.objects.get_or_create(
                    curso=curso, nombre=seccion_nombre, defaults=seccion_defaults
                )
                if s_created:
                    created_secciones += 1
                elif profesor and seccion.profesor_id != profesor.id:
                    seccion.profesor = profesor
                    seccion.save(update_fields=["profesor"])

                # ----- Estudiante (buscar por código o por email) -----
                est = Estudiante.objects.filter(codigo=codigo).first()
                if not est:
                    est = Estudiante.objects.filter(email=email).first()

                if est:
                    updates = []
                    if est.nombre != nombre:
                        est.nombre = nombre; updates.append("nombre")
                    if est.apellido != apellido:
                        est.apellido = apellido; updates.append("apellido")
                    if est.email != email:
                        # evita colisión de email único
                        if not Estudiante.objects.filter(email=email).exclude(pk=est.pk).exists():
                            est.email = email; updates.append("email")
                        else:
                            self.stderr.write(self.style.WARNING(
                                f"⚠ Email duplicado '{email}' ya usado por otro estudiante. Mantengo el existente para codigo={est.codigo}."
                            ))
                    if est.codigo != codigo:
                        if not Estudiante.objects.filter(codigo=codigo).exclude(pk=est.pk).exists():
                            est.codigo = codigo; updates.append("codigo")
                        else:
                            self.stderr.write(self.style.WARNING(
                                f"⚠ Código duplicado '{codigo}' ya usado por otro estudiante. Mantengo el existente para email={est.email}."
                            ))
                    if updates:
                        est.save(update_fields=updates)
                else:
                    try:
                        est = Estudiante.objects.create(
                            codigo=codigo, nombre=nombre, apellido=apellido, email=email
                        )
                        created_estudiantes += 1
                    except IntegrityError:
                        # Intento de rescate por colisión de email/código
                        est = Estudiante.objects.filter(email=email).first() or Estudiante.objects.filter(codigo=codigo).first()
                        if not est:
                            raise  # re-lanzar si de verdad no hay forma de resolver
                        updates = []
                        if est.nombre != nombre: est.nombre = nombre; updates.append("nombre")
                        if est.apellido != apellido: est.apellido = apellido; updates.append("apellido")
                        if est.codigo != codigo and not Estudiante.objects.filter(codigo=codigo).exclude(pk=est.pk).exists():
                            est.codigo = codigo; updates.append("codigo")
                        if updates: est.save(update_fields=updates)

                # ----- Nota semilla -----
                if crear_nota:
                    _, n_created = Nota.objects.get_or_create(
                        seccion=seccion, estudiante=est,
                        defaults=dict(avance1=0, avance2=0, avance3=0,
                                      participacion=0, proyecto_final=0, nota_final=0)
                    )
                    if n_created:
                        created_notas += 1

                total_rows += 1

        self.stdout.write(self.style.SUCCESS(
            f"✔ Importado OK: filas={total_rows}, cursos_nuevos={created_cursos}, "
            f"secciones_nuevas={created_secciones}, estudiantes_nuevos={created_estudiantes}, "
            f"notas_creadas={created_notas}"
        ))
